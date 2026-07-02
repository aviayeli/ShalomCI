from openpyxl import load_workbook

from src.core.reporter import ExcelReporter


def test_excel_report_generation(tmp_path):
    """מוודא יצירת אקסל תקינה עם קידודי צבעים לפי ציון (אדום/צהוב/ירוק)."""
    reporter = ExcelReporter()

    test_data = [
        {"mpn": "PART_EOL", "lifecycle_status": "Obsolete", "risk_score": 1, "alternatives": []},
        {"mpn": "PART_NRND", "lifecycle_status": "NRND", "risk_score": 3, "alternatives": [{"mpn": "ALT1"}]},
        {"mpn": "PART_GOOD", "lifecycle_status": "Active", "risk_score": 5, "alternatives": []}
    ]

    file_path = tmp_path / "test_report.xlsx"
    reporter.generate_report(test_data, str(file_path))

    # נטען את הקובץ מחדש כדי לוודא עיצובים
    wb = load_workbook(str(file_path))
    ws = wb.active

    assert ws.cell(row=1, column=1).value == 'מק"ט יצרן (MPN)'

    # בדיקת צבע אדום לציון 1
    assert ws.cell(row=2, column=4).fill.start_color.rgb == "00FFCCCC"
    # בדיקת חלופה כסטרינג
    assert ws.cell(row=3, column=5).value == "ALT1"
    # בדיקת צבע ירוק לציון 5
    assert ws.cell(row=4, column=4).fill.start_color.rgb == "00CCFFCC"
