from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class ExcelReporter:
    """מנוע להפקת דוחות אקסל בעברית עם קידוד צבעים לפי רמות סיכון."""

    def __init__(self):
        # הגדרת צבעים להארת תאים באקסל
        self.color_critical = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.color_warning = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        self.color_healthy = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    def generate_report(self, components: List[Dict[str, Any]], output_path: str):
        """מייצר ושומר את הדוח לקובץ."""
        wb = Workbook()
        ws = wb.active
        ws.title = "סיכום רכיבים - ShalomCI"

        # הגדרת הגיליון כולו מימין לשמאל (RTL)
        ws.sheet_view.rightToLeft = True

        # כותרות בעברית
        headers = ["מק\"ט יצרן (MPN)", "יצרן", "סטטוס מחזור חיים", "ציון סיכון", "חלופות מוצעות"]
        ws.append(headers)

        # הדגשת כותרות
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        for row_idx, comp in enumerate(components, start=2):
            alts = comp.get("alternatives", [])
            alt_str = ", ".join([a.get("mpn", "Unknown") for a in alts]) if alts else "אין חלופות"

            ws.cell(row=row_idx, column=1, value=comp.get("mpn", ""))
            ws.cell(row=row_idx, column=2, value=comp.get("manufacturer", ""))
            ws.cell(row=row_idx, column=3, value=comp.get("lifecycle_status", "לא ידוע"))

            score = comp.get("risk_score", 0)
            score_cell = ws.cell(row=row_idx, column=4, value=score)
            ws.cell(row=row_idx, column=5, value=alt_str)

            # צביעת תא הסיכון
            if score == 1:
                score_cell.fill = self.color_critical
            elif score in (2, 3):
                score_cell.fill = self.color_warning
            elif score in (4, 5):
                score_cell.fill = self.color_healthy

        wb.save(output_path)